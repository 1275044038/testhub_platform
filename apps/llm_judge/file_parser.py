"""批量评分文件解析：支持 CSV / XLSX / TXT。

返回统一结构：
{
  "filename": str,
  "total_rows": int,
  "valid_rows": int,
  "errors": [str],
  "preview": [dict],           # 前 10 条，用于前端表格预览
  "cases": [dict],             # 完整用例列表（字段：question, answer, ground_truth?, auto_gt?, context?）
}

格式约定：
  - CSV/XLSX：表头需包含「问题/question」「答案/answer」两列；可选列：参考/ground_truth、自动GT/auto_gt、上下文/context
  - TXT：每行一条，格式同前端批量输入：「问题 ||| 答案」，每行可选 ||| 参考(第三段)
"""
from __future__ import annotations

import csv
import io
import json
import logging
from typing import Any

logger = logging.getLogger(__name__)

# 表头别名映射
QUESTION_COLS = {'问题', 'question', 'Question', 'QUESTION', 'query', 'Query', 'prompt'}
ANSWER_COLS = {'答案', '回答', 'answer', 'Answer', 'ANSWER', 'response', 'Response'}
GT_COLS = {'参考', '参考答案', 'ground_truth', 'gt', 'GT', 'GroundTruth', 'ground truth'}
AUTO_GT_COLS = {'自动匹配', '自动GT', 'auto_gt', 'auto gt', 'autoGt', '自动参考', '自动匹配参考答案', '自动匹配gt', '是否自动匹配参考'}
CONTEXT_COLS = {'上下文', 'context', 'Context', '背景'}

MAX_CASES = 5000
MAX_PREVIEW = 10


def _norm_header(h: Any) -> str:
    return str(h or '').strip()


def _to_bool(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    s = str(v or '').strip().lower()
    return s in {'1', 'true', 'yes', 'y', '是', '对', 't'}


def _to_json_gt(v: Any):
    if v is None or v == '':
        return None
    s = str(v).strip()
    if not s:
        return None
    if s.startswith('{') or s.startswith('['):
        try:
            return json.loads(s)
        except Exception:
            return {'text': s}
    return {'text': s}


def parse_text(content: str, filename: str = '') -> dict:
    errors: list[str] = []
    cases: list[dict] = []
    lines = (content or '').splitlines()
    for i, raw in enumerate(lines, 1):
        line = raw.strip()
        if not line:
            continue
        parts = line.split('|||')
        if len(parts) < 2:
            # 兼容 tab / 3个以上竖线
            parts = [p.strip() for p in line.split('\t')]
        if len(parts) < 2:
            errors.append(f'第{i}行格式错误：需要「问题 ||| 答案」，实际: {line[:40]}')
            continue
        question = parts[0].strip()
        answer = parts[1].strip()
        if not question or not answer:
            errors.append(f'第{i}行：问题或答案为空')
            continue
        case = {'question': question, 'answer': answer, 'auto_gt': True}
        if len(parts) >= 3 and parts[2].strip():
            gt_text = parts[2].strip()
            case['ground_truth'] = {'text': gt_text}
            case['auto_gt'] = False
        cases.append(case)

    total = len(lines)
    valid = len(cases)
    if valid > MAX_CASES:
        errors.append(f'用例数超过上限 {MAX_CASES}，已截取前 {MAX_CASES} 条')
        cases = cases[:MAX_CASES]
        valid = MAX_CASES
    return {
        'filename': filename,
        'total_rows': total,
        'valid_rows': valid,
        'errors': errors,
        'preview': cases[:MAX_PREVIEW],
        'cases': cases,
    }


def parse_csv(content_bytes: bytes, filename: str = '') -> dict:
    errors: list[str] = []
    cases: list[dict] = []
    total_rows = 0
    try:
        text = content_bytes.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = content_bytes.decode('gbk', errors='ignore')
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return {'filename': filename, 'total_rows': 0, 'valid_rows': 0,
                'errors': ['文件为空'], 'preview': [], 'cases': []}

    header = [_norm_header(h) for h in rows[0]]
    # 找列索引
    q_idx = a_idx = gt_idx = auto_idx = ctx_idx = -1
    for i, h in enumerate(header):
        if q_idx < 0 and h in QUESTION_COLS:
            q_idx = i
        elif a_idx < 0 and h in ANSWER_COLS:
            a_idx = i
        elif gt_idx < 0 and h in GT_COLS:
            gt_idx = i
        elif auto_idx < 0 and h in AUTO_GT_COLS:
            auto_idx = i
        elif ctx_idx < 0 and h in CONTEXT_COLS:
            ctx_idx = i

    if q_idx < 0 or a_idx < 0:
        # 无表头：按 问题,答案,GT 三列约定
        if len(header) >= 2:
            q_idx, a_idx = 0, 1
            gt_idx = 2 if len(header) > 2 else -1
            # 首行作为数据回退
            rows.insert(0, None)  # placeholder，不删首行
            rows.pop(0)  # 原首行就是 header，不回退
            # 重新判断：如果 header 第一个不是「问题/question」关键字，把第一行当数据
            if not (header[0] in QUESTION_COLS or header[1] in ANSWER_COLS):
                data_rows = rows  # header 当作数据
                # 重新索引
                q_idx, a_idx, gt_idx = 0, 1, (2 if len(header) > 2 else -1)
                auto_idx = ctx_idx = -1
            else:
                data_rows = rows[1:]
        else:
            return {'filename': filename, 'total_rows': len(rows), 'valid_rows': 0,
                    'errors': ['CSV 表头缺少「问题」和「答案」列'], 'preview': [], 'cases': []}
    else:
        data_rows = rows[1:]

    for i, row in enumerate(data_rows, 2):
        total_rows += 1
        if not row or all((str(c).strip() == '' for c in row)):
            continue
        def _get(idx):
            return row[idx] if 0 <= idx < len(row) else ''
        question = str(_get(q_idx)).strip()
        answer = str(_get(a_idx)).strip()
        if not question or not answer:
            errors.append(f'第{i}行：问题或答案为空')
            continue
        case = {'question': question, 'answer': answer}
        if gt_idx >= 0:
            gt = _to_json_gt(_get(gt_idx))
            if gt:
                case['ground_truth'] = gt
                case['auto_gt'] = False
            else:
                case['auto_gt'] = True
        else:
            case['auto_gt'] = True
        # auto 列优先：显式指定后覆盖默认推断；auto=False 且无 GT 时给出错误
        if auto_idx >= 0:
            auto_v = _to_bool(_get(auto_idx))
            if auto_v:
                case['auto_gt'] = True
                case.pop('ground_truth', None)
            else:
                case['auto_gt'] = False
                if not case.get('ground_truth'):
                    errors.append(f'第{i}行：已标记为手动 GT（自动匹配=否）但参考答案为空')
                    continue
        if ctx_idx >= 0:
            ctx_raw = str(_get(ctx_idx)).strip()
            if ctx_raw:
                try:
                    case['context'] = json.loads(ctx_raw)
                except Exception:
                    case['context'] = {'text': ctx_raw}
        cases.append(case)
        if len(cases) >= MAX_CASES:
            break

    valid = len(cases)
    if valid >= MAX_CASES and total_rows > MAX_CASES:
        errors.append(f'用例数超过上限 {MAX_CASES}，已截取前 {MAX_CASES} 条')
    return {
        'filename': filename,
        'total_rows': total_rows,
        'valid_rows': valid,
        'errors': errors,
        'preview': cases[:MAX_PREVIEW],
        'cases': cases,
    }


def parse_xlsx(content_bytes: bytes, filename: str = '') -> dict:
    try:
        import openpyxl
    except ImportError:
        return {'filename': filename, 'total_rows': 0, 'valid_rows': 0,
                'errors': ['缺少 openpyxl 依赖，无法解析 XLSX。请安装: pip install openpyxl'],
                'preview': [], 'cases': []}
    errors: list[str] = []
    cases: list[dict] = []
    total_rows = 0
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return {'filename': filename, 'total_rows': 0, 'valid_rows': 0,
                'errors': [f'XLSX 解析失败: {e}'], 'preview': [], 'cases': []}

    if not rows:
        return {'filename': filename, 'total_rows': 0, 'valid_rows': 0,
                'errors': ['文件为空'], 'preview': [], 'cases': []}

    header = [_norm_header(h) for h in rows[0]]
    q_idx = a_idx = gt_idx = auto_idx = ctx_idx = -1
    for i, h in enumerate(header):
        if q_idx < 0 and h in QUESTION_COLS:
            q_idx = i
        elif a_idx < 0 and h in ANSWER_COLS:
            a_idx = i
        elif gt_idx < 0 and h in GT_COLS:
            gt_idx = i
        elif auto_idx < 0 and h in AUTO_GT_COLS:
            auto_idx = i
        elif ctx_idx < 0 and h in CONTEXT_COLS:
            ctx_idx = i
    if q_idx < 0 or a_idx < 0:
        return {'filename': filename, 'total_rows': len(rows), 'valid_rows': 0,
                'errors': ['XLSX 表头缺少「问题」和「答案」列'], 'preview': [], 'cases': []}

    for i, row in enumerate(rows[1:], 2):
        total_rows += 1
        if row is None or all((c is None or str(c).strip() == '' for c in row)):
            continue
        def _get(idx):
            return row[idx] if 0 <= idx < len(row) else ''
        question = str(_get(q_idx)).strip()
        answer = str(_get(a_idx)).strip()
        if not question or not answer:
            errors.append(f'第{i}行：问题或答案为空')
            continue
        case = {'question': question, 'answer': answer}
        if gt_idx >= 0:
            gt = _to_json_gt(_get(gt_idx))
            if gt:
                case['ground_truth'] = gt
                case['auto_gt'] = False
            else:
                case['auto_gt'] = True
        else:
            case['auto_gt'] = True
        # auto 列优先：显式指定后覆盖默认推断；auto=False 且无 GT 时给出错误
        if auto_idx >= 0:
            auto_v = _to_bool(_get(auto_idx))
            if auto_v:
                case['auto_gt'] = True
                case.pop('ground_truth', None)
            else:
                case['auto_gt'] = False
                if not case.get('ground_truth'):
                    errors.append(f'第{i}行：已标记为手动 GT（自动匹配=否）但参考答案为空')
                    continue
        if ctx_idx >= 0:
            ctx_raw = str(_get(ctx_idx)).strip()
            if ctx_raw:
                try:
                    case['context'] = json.loads(ctx_raw)
                except Exception:
                    case['context'] = {'text': ctx_raw}
        cases.append(case)
        if len(cases) >= MAX_CASES:
            break

    valid = len(cases)
    if valid >= MAX_CASES and total_rows > MAX_CASES:
        errors.append(f'用例数超过上限 {MAX_CASES}，已截取前 {MAX_CASES} 条')
    return {
        'filename': filename,
        'total_rows': total_rows,
        'valid_rows': valid,
        'errors': errors,
        'preview': cases[:MAX_PREVIEW],
        'cases': cases,
    }


def detect_and_parse(filename: str, content_bytes: bytes | str) -> dict:
    """统一入口：根据文件名后缀选择解析器。外层 try/except 确保任何异常都返回友好错误。"""
    try:
        fn = (filename or '').lower()
        if isinstance(content_bytes, str):
            return parse_text(content_bytes, filename)
        # 文件大小先校验
        if isinstance(content_bytes, (bytes, bytearray)) and len(content_bytes) > 10 * 1024 * 1024:
            return {'filename': filename, 'total_rows': 0, 'valid_rows': 0,
                    'errors': [f'文件超过 10MB 上限（当前 {len(content_bytes) // 1024} KB）'], 'preview': [], 'cases': []}
        if fn.endswith('.csv'):
            return parse_csv(content_bytes, filename)
        if fn.endswith('.xlsx') or fn.endswith('.xlsm'):
            return parse_xlsx(content_bytes, filename)
        if fn.endswith('.txt') or fn.endswith('.md') or not fn:
            try:
                text = content_bytes.decode('utf-8-sig')
            except UnicodeDecodeError:
                text = content_bytes.decode('gbk', errors='ignore')
            return parse_text(text, filename)
        try:
            text = content_bytes.decode('utf-8-sig')
        except Exception:
            text = content_bytes.decode('gbk', errors='ignore')
        return parse_text(text, filename)
    except Exception as exc:
        return {'filename': filename, 'total_rows': 0, 'valid_rows': 0,
                'errors': [f'文件解析失败：{exc}'], 'preview': [], 'cases': []}


def build_template_bytes(kind: str = 'csv') -> tuple[bytes, str, str]:
    """生成批量用例模板文件。返回 (bytes, filename, content_type)。"""
    header = ['问题', '答案', '参考答案', '自动匹配参考答案', '上下文']
    sample_rows = [
        ['贵州茅台2024年报营业收入是多少？', '根据公司年报，贵州茅台2024年实现营业收入约1741.4亿元。', '', '是', ''],
        ['比亚迪2024年销量多少？', '比亚迪2024年销量达到427.2万辆。', '', '是', ''],
        ['什么是机器学习？', '机器学习是人工智能的一个分支，通过数据训练模型进行预测或决策。', '机器学习是一种利用数据进行模式识别的AI方法', '否', '{"领域":"AI"}'],
    ]
    if kind == 'csv':
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(header)
        for r in sample_rows:
            w.writerow(r)
        data = buf.getvalue().encode('utf-8-sig')
        return data, 'judge_batch_template.csv', 'text/csv; charset=utf-8'
    if kind == 'txt':
        lines = []
        for r in sample_rows:
            q, a, gt, auto, ctx = r
            parts = [q, a]
            if gt and auto != '是':
                parts.append(gt)
            lines.append(' ||| '.join(parts))
        data = ('# 每行一条，格式：问题 ||| 答案 [ ||| 参考答案(可选)]\n' +
                '\n'.join(lines) + '\n').encode('utf-8')
        return data, 'judge_batch_template.txt', 'text/plain; charset=utf-8'
    if kind == 'xlsx':
        try:
            import openpyxl
        except ImportError:
            # fallback csv
            return build_template_bytes('csv')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '批量用例'
        ws.append(header)
        for r in sample_rows:
            ws.append(r)
        # 列宽
        for col, w in zip('ABCDE', [40, 50, 40, 20, 20]):
            ws.column_dimensions[col].width = w
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), 'judge_batch_template.xlsx', \
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return build_template_bytes('csv')
